from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.contrib.auth import update_session_auth_hash
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.utils import timezone
from django.core.paginator import Paginator
from django.db import IntegrityError
from .models import Teacher, Student, Attendance, TeacherAttendance
from .sms_utils import build_absent_message, build_teacher_absent_message, send_sms
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# Expected Excel column headers (in order) for the two bulk-upload forms.
# These must exactly match (case/whitespace-insensitive) the header row
# shown to the admin in student_upload.html / teacher_upload.html.
STUDENT_EXCEL_HEADERS = ["Roll", "Name", "Class", "Section", "Session", "Phone"]
TEACHER_EXCEL_HEADERS = ["Name", "Number", "Class"]


def is_admin(user):
    return user.is_staff


def check_header(header_row, expected_headers):
    """
    Compares the first row of an uploaded Excel file against the expected
    column headers (case/whitespace-insensitive, order matters).
    Returns None if it matches, otherwise a human-readable error message.
    """
    cells = list(header_row) if header_row else []
    actual = []
    for i in range(len(expected_headers)):
        cell = cells[i] if i < len(cells) else None
        actual.append(str(cell).strip().lower() if cell is not None else "")
    expected = [h.strip().lower() for h in expected_headers]

    if actual != expected:
        expected_display = ", ".join(expected_headers)
        return (
            f"Header row is not in the correct format. The first row must have "
            f"these exact columns, in this order: {expected_display}. "
            f"Please fix the header and upload again."
        )
    return None


def get_class_choices():
    import re

    names = set()
    for c in Student.objects.values_list('class_name', flat=True):
        if c is None:
            continue
        cleaned = str(c).strip()
        if cleaned:
            names.add(cleaned)

    def sort_key(name):
        s = str(name).strip()
        low = s.lower()

        if low.startswith(('play', 'nursery', 'kg', 'pre')):
            group = 0
        elif any(ch.isdigit() for ch in s):
            group = 1
        else:
            group = 2

        m = re.search(r'(\d+)', s)
        num = int(m.group(1)) if m else 0
        return (group, num, low)

    return sorted(names, key=sort_key)


def get_section_choices():
    return list(
        Student.objects.exclude(section='').values_list('section', flat=True).distinct().order_by('section')
    )


def build_choice_options(values, selected_value):
    return [{'value': v, 'is_selected': (str(v) == str(selected_value))} for v in values]


def build_choice_options_multi(values, selected_values):
    selected_set = {str(v) for v in selected_values}
    return [{'value': v, 'is_selected': (str(v) in selected_set)} for v in values]


@login_required
def dashboard(request):
    teacher = Teacher.objects.filter(user=request.user).first()
    total_students = 0
    present_count = 0
    absent_count = 0
    today = timezone.now().date()
    teacher_classes = []

    if teacher:
        teacher_classes = teacher.get_class_list()
        students = Student.objects.filter(class_name__in=teacher_classes)
        total_students = students.count()
        attendance_today = Attendance.objects.filter(
            student__class_name__in=teacher_classes,
            date=today
        )
        present_count = attendance_today.filter(is_present=True).count()
        absent_count = attendance_today.filter(is_present=False).count()

    if request.user.is_staff:
        all_today = Attendance.objects.filter(date=today)
        admin_present = all_today.filter(is_present=True).count()
        admin_absent = all_today.filter(is_present=False).count()
        admin_total = Student.objects.count()
    else:
        admin_present = admin_absent = admin_total = 0

    return render(request, 'attendance/dashboard.html', {
        'teacher': teacher,
        'teacher_classes': teacher_classes,
        'total_students': total_students,
        'present_count': present_count,
        'absent_count': absent_count,
        'today': today,
        'admin_present': admin_present,
        'admin_absent': admin_absent,
        'admin_total': admin_total,
    })


@login_required
def change_password(request):
    error = None
    success = False

    if request.method == 'POST':
        old_password = request.POST.get('old_password', '')
        new_password = request.POST.get('new_password', '')
        confirm_password = request.POST.get('confirm_password', '')

        if not request.user.check_password(old_password):
            error = "Current password is incorrect."
        elif len(new_password) < 4:
            error = "New password must be at least 4 characters."
        elif new_password != confirm_password:
            error = "New passwords do not match."
        else:
            request.user.set_password(new_password)
            request.user.save()
            update_session_auth_hash(request, request.user)
            success = True

    return render(request, 'attendance/change_password.html', {'error': error, 'success': success})


@login_required
def mark_attendance(request):
    teacher = Teacher.objects.filter(user=request.user).first()
    is_admin_user = request.user.is_staff
    selected_class = request.GET.get('class', '').strip()

    if not is_admin_user and not teacher:
        return render(request, 'attendance/mark_attendance.html', {
            'error': 'Your account is not linked to any teacher. Ask admin to link it.',
            'is_admin_user': False,
        })

    if is_admin_user:
        class_choices = get_class_choices()
    else:
        class_choices = teacher.get_class_list()
        if not class_choices:
            return render(request, 'attendance/mark_attendance.html', {
                'error': 'No class has been assigned to you yet. Please ask the admin to assign one.',
                'is_admin_user': False,
            })

    show_class_selector = is_admin_user or len(class_choices) > 1

    if not is_admin_user and len(class_choices) == 1:
        current_class = class_choices[0]
    elif selected_class in class_choices:
        current_class = selected_class
    else:
        current_class = None

    all_classes = [{'name': c, 'is_selected': (c == current_class)} for c in class_choices]

    def load_students(class_name):
        if is_admin_user:
            return Student.objects.filter(class_name=class_name).order_by('section', 'roll_no')
        return Student.objects.filter(class_name=class_name).order_by('roll_no')

    students = load_students(current_class) if current_class else Student.objects.none()

    today = timezone.now().date()
    date_str = today.strftime("%d-%b-%y")

    already_marked = False
    if current_class:
        already_marked = Attendance.objects.filter(
            student__class_name=current_class,
            date=today
        ).exists()

    saved = False
    sms_sent_count = 0
    sms_warning = None

    if request.method == 'POST' and current_class and not already_marked:
        post_class = request.POST.get('class', '').strip()
        if post_class and post_class in class_choices and post_class != current_class:
            current_class = post_class
            students = load_students(current_class)
            all_classes = [{'name': c, 'is_selected': (c == current_class)} for c in class_choices]
            already_marked = Attendance.objects.filter(
                student__class_name=current_class, date=today
            ).exists()

        if not already_marked:
            sms_failed = []
            for student in students:
                status = request.POST.get(f'att_{student.id}', 'present')
                is_present = status == 'present'

                Attendance.objects.update_or_create(
                    student=student,
                    date=today,
                    defaults={'is_present': is_present}
                )

                if not is_present:
                    if student.parent_mobile:
                        message = build_absent_message(student.name, date_str)
                        success, resp = send_sms(student.parent_mobile, message)
                        if success:
                            sms_sent_count += 1
                        else:
                            sms_failed.append(student.name)
                    else:
                        sms_failed.append(f"{student.name} (no phone)")

            saved = True
            already_marked = True
            if sms_failed:
                sms_warning = f"SMS could not be sent to: {', '.join(sms_failed)}"

    attendance_map = {}
    if current_class:
        attendance_map = {
            a.student_id: a.is_present
            for a in Attendance.objects.filter(student__class_name=current_class, date=today)
        }

    student_rows = [
        {
            'student': s,
            'is_present': attendance_map.get(s.id, True),
        }
        for s in students
    ]

    if already_marked and current_class:
        present_count = Attendance.objects.filter(
            student__class_name=current_class, date=today, is_present=True
        ).count()
        absent_count = Attendance.objects.filter(
            student__class_name=current_class, date=today, is_present=False
        ).count()
    else:
        present_count = 0
        absent_count = 0

    return render(request, 'attendance/mark_attendance.html', {
        'teacher': teacher,
        'students': students,
        'student_rows': student_rows,
        'today': today,
        'saved': saved,
        'sms_sent_count': sms_sent_count,
        'sms_warning': sms_warning,
        'is_admin_user': is_admin_user,
        'show_class_selector': show_class_selector,
        'all_classes': all_classes,
        'current_class': current_class,
        'already_marked': already_marked,
        'present_count': present_count,
        'absent_count': absent_count,
        'total_students': students.count() if hasattr(students, 'count') else len(students),
    })


@login_required
@user_passes_test(is_admin)
def student_list(request):
    query = request.GET.get('q', '').strip()
    class_filter = request.GET.get('class', '').strip()

    class_names = get_class_choices()
    all_classes = [{'name': c, 'is_selected': (str(c) == class_filter)} for c in class_names]

    students = Student.objects.all().order_by('class_name', 'section', 'roll_no')
    if class_filter:
        students = students.filter(class_name=class_filter)
    if query:
        students = students.filter(name__icontains=query)

    total_count = students.count()

    paginator = Paginator(students, 50)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    return render(request, 'attendance/student_list.html', {
        'students': page_obj,
        'total_count': total_count,
        'query': query,
        'class_filter': class_filter,
        'all_classes': all_classes,
    })


@login_required
@user_passes_test(is_admin)
def student_add(request):
    error = None
    if request.method == 'POST':
        try:
            Student.objects.create(
                roll_no=request.POST.get('roll_no', '').strip(),
                name=request.POST.get('name', '').strip(),
                class_name=request.POST.get('class_name', '').strip(),
                section=request.POST.get('section', '').strip(),
                parent_mobile=request.POST.get('parent_mobile', '').strip(),
            )
            return redirect('student_list')
        except IntegrityError:
            error = "A student with this Roll Number already exists in this Class/Section. Please check and try again."

    return render(request, 'attendance/student_form.html', {
        'mode': 'Add',
        'error': error,
        'class_options': build_choice_options(get_class_choices(), ''),
        'section_options': build_choice_options(get_section_choices(), ''),
    })


@login_required
@user_passes_test(is_admin)
def student_edit(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    error = None
    if request.method == 'POST':
        try:
            student.roll_no = request.POST.get('roll_no', '').strip()
            student.name = request.POST.get('name', '').strip()
            student.class_name = request.POST.get('class_name', '').strip()
            student.section = request.POST.get('section', '').strip()
            student.parent_mobile = request.POST.get('parent_mobile', '').strip()
            student.save()
            return redirect('student_list')
        except IntegrityError:
            error = "A student with this Roll Number already exists in this Class/Section. Please check and try again."

    return render(request, 'attendance/student_form.html', {
        'mode': 'Edit',
        'student': student,
        'error': error,
        'class_options': build_choice_options(get_class_choices(), student.class_name),
        'section_options': build_choice_options(get_section_choices(), student.section),
    })


@login_required
@user_passes_test(is_admin)
def student_delete(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    if request.method == 'POST':
        student.delete()
        return redirect('student_list')
    return render(request, 'attendance/student_confirm_delete.html', {'student': student})


@login_required
@user_passes_test(is_admin)
def class_delete(request, class_name):
    students = Student.objects.filter(class_name=class_name)
    count = students.count()

    if request.method == 'POST':
        students.delete()
        return redirect('student_list')

    return render(request, 'attendance/class_confirm_delete.html', {
        'class_name': class_name,
        'count': count,
    })


@login_required
@user_passes_test(is_admin)
def student_upload(request):
    file_results = []

    if request.method == 'POST' and request.FILES.getlist('excel_file'):
        for excel_file in request.FILES.getlist('excel_file'):

            if not excel_file.name.lower().endswith('.xlsx'):
                file_results.append({
                    "filename": excel_file.name,
                    "error": "This is not an .xlsx file. Please save it as an Excel (.xlsx) file and try again.",
                })
                continue

            try:
                wb = openpyxl.load_workbook(excel_file, data_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
            except Exception:
                file_results.append({
                    "filename": excel_file.name,
                    "error": "Could not open this file. It may be corrupted, password-protected, or not a real Excel file.",
                })
                continue

            if not rows:
                file_results.append({
                    "filename": excel_file.name,
                    "error": "This file appears to be empty (only a header row, or no rows at all).",
                })
                continue

            header_error = check_header(rows[0], STUDENT_EXCEL_HEADERS)
            if header_error:
                file_results.append({
                    "filename": excel_file.name,
                    "error": header_error,
                })
                continue

            if len(rows) < 2:
                file_results.append({
                    "filename": excel_file.name,
                    "error": "This file appears to be empty (only a header row, or no rows at all).",
                })
                continue

            created, updated, skipped = 0, 0, 0
            detected_classes = set()

            for row in rows[1:]:
                if not row or len(row) < 3:
                    skipped += 1
                    continue

                roll = row[0]
                name = row[1]
                class_name = row[2]
                section = row[3] or "" if len(row) > 3 else ""
                phone = row[5] if len(row) > 5 else None

                if not roll or not name or not class_name:
                    skipped += 1
                    continue

                clean_phone = str(phone).strip() if phone else ""
                if clean_phone and not clean_phone.startswith("0"):
                    clean_phone = "0" + clean_phone

                obj, was_created = Student.objects.update_or_create(
                    class_name=str(class_name).strip(),
                    section=str(section).strip(),
                    roll_no=str(roll).strip(),
                    defaults={
                        "name": str(name).strip(),
                        "parent_mobile": clean_phone,
                    }
                )
                detected_classes.add(f"{class_name}{section}")
                if was_created:
                    created += 1
                else:
                    updated += 1

            if created == 0 and updated == 0:
                file_results.append({
                    "filename": excel_file.name,
                    "error": f"No valid rows found. All {skipped} row(s) were missing a Roll, Name, or Class value.",
                })
                continue

            file_results.append({
                "filename": excel_file.name,
                "class_label": ", ".join(sorted(detected_classes)) or "Unknown",
                "created": created,
                "updated": updated,
                "skipped": skipped,
            })

    return render(request, 'attendance/student_upload.html', {'file_results': file_results})


@login_required
@user_passes_test(is_admin)
def teacher_list(request):
    full_time = Teacher.objects.select_related('user').filter(
        employment_type=Teacher.EMPLOYMENT_FULL
    ).order_by('id')
    part_time = Teacher.objects.select_related('user').filter(
        employment_type=Teacher.EMPLOYMENT_PART
    ).order_by('id')
    return render(request, 'attendance/teacher_list.html', {
        'full_time_teachers': full_time,
        'part_time_teachers': part_time,
    })


@login_required
@user_passes_test(is_admin)
def teacher_add(request):
    error = None
    employment_type = request.GET.get('type', request.POST.get('employment_type', Teacher.EMPLOYMENT_FULL)).strip()
    if employment_type not in (Teacher.EMPLOYMENT_FULL, Teacher.EMPLOYMENT_PART):
        employment_type = Teacher.EMPLOYMENT_FULL

    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '').strip()
        name = request.POST.get('name', '').strip()
        mobile = request.POST.get('mobile', '').strip()
        employment_type = request.POST.get('employment_type', Teacher.EMPLOYMENT_FULL).strip()
        if employment_type not in (Teacher.EMPLOYMENT_FULL, Teacher.EMPLOYMENT_PART):
            employment_type = Teacher.EMPLOYMENT_FULL
        assigned_classes_list = request.POST.getlist('assigned_classes')
        assigned_classes = ",".join(c.strip() for c in assigned_classes_list if c.strip())

        if User.objects.filter(username=username).exists():
            error = f"Username '{username}' is already taken."
        else:
            user = User.objects.create_user(username=username, password=password)
            Teacher.objects.create(
                user=user,
                name=name,
                mobile=mobile,
                assigned_classes=assigned_classes,
                employment_type=employment_type,
            )
            return redirect('teacher_list')

    type_label = 'Full-time' if employment_type == Teacher.EMPLOYMENT_FULL else 'Part-time'
    return render(request, 'attendance/teacher_form.html', {
        'mode': 'Add',
        'error': error,
        'class_options': build_choice_options_multi(get_class_choices(), []),
        'employment_type': employment_type,
        'type_label': type_label,
    })


@login_required
@user_passes_test(is_admin)
def teacher_edit(request, teacher_id):
    teacher = get_object_or_404(Teacher, id=teacher_id)
    if request.method == 'POST':
        teacher.name = request.POST.get('name', '').strip()
        teacher.mobile = request.POST.get('mobile', '').strip()
        employment_type = request.POST.get('employment_type', Teacher.EMPLOYMENT_FULL).strip()
        if employment_type not in (Teacher.EMPLOYMENT_FULL, Teacher.EMPLOYMENT_PART):
            employment_type = Teacher.EMPLOYMENT_FULL
        teacher.employment_type = employment_type
        assigned_classes_list = request.POST.getlist('assigned_classes')
        teacher.assigned_classes = ",".join(c.strip() for c in assigned_classes_list if c.strip())
        teacher.save()

        new_password = request.POST.get('password', '').strip()
        if new_password and teacher.user:
            teacher.user.set_password(new_password)
            teacher.user.save()

        return redirect('teacher_list')

    type_label = 'Full-time' if teacher.employment_type == Teacher.EMPLOYMENT_FULL else 'Part-time'
    return render(request, 'attendance/teacher_form.html', {
        'mode': 'Edit',
        'teacher': teacher,
        'class_options': build_choice_options_multi(get_class_choices(), teacher.get_class_list()),
        'employment_type': teacher.employment_type,
        'type_label': type_label,
    })


@login_required
@user_passes_test(is_admin)
def teacher_upload(request):
    file_results = []
    employment_type = request.GET.get('type', request.POST.get('employment_type', Teacher.EMPLOYMENT_FULL)).strip()
    if employment_type not in (Teacher.EMPLOYMENT_FULL, Teacher.EMPLOYMENT_PART):
        employment_type = Teacher.EMPLOYMENT_FULL
    type_label = 'Full-time' if employment_type == Teacher.EMPLOYMENT_FULL else 'Part-time'

    if request.method == 'POST' and request.FILES.getlist('excel_file'):
        employment_type = request.POST.get('employment_type', Teacher.EMPLOYMENT_FULL).strip()
        if employment_type not in (Teacher.EMPLOYMENT_FULL, Teacher.EMPLOYMENT_PART):
            employment_type = Teacher.EMPLOYMENT_FULL
        type_label = 'Full-time' if employment_type == Teacher.EMPLOYMENT_FULL else 'Part-time'

        for excel_file in request.FILES.getlist('excel_file'):
            if not excel_file.name.lower().endswith('.xlsx'):
                file_results.append({
                    "filename": excel_file.name,
                    "error": "Only .xlsx files are allowed.",
                })
                continue

            try:
                wb = openpyxl.load_workbook(excel_file, data_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
            except Exception:
                file_results.append({
                    "filename": excel_file.name,
                    "error": "Could not open this file.",
                })
                continue

            if not rows:
                file_results.append({
                    "filename": excel_file.name,
                    "error": "File is empty.",
                })
                continue

            header_error = check_header(rows[0], TEACHER_EXCEL_HEADERS)
            if header_error:
                file_results.append({
                    "filename": excel_file.name,
                    "error": header_error,
                })
                continue

            if len(rows) < 2:
                file_results.append({
                    "filename": excel_file.name,
                    "error": "This file appears to be empty (only a header row, no data rows).",
                })
                continue

            created, updated, skipped = 0, 0, 0

            for row in rows[1:]:
                if not row or len(row) < 1:
                    skipped += 1
                    continue

                name = row[0]
                mobile = row[1] if len(row) > 1 else ""
                assigned_class = row[2] if len(row) > 2 else ""

                if not name:
                    skipped += 1
                    continue

                clean_mobile = str(mobile).strip() if mobile else ""
                if clean_mobile.endswith('.0'):
                    clean_mobile = clean_mobile[:-2]
                if clean_mobile and not clean_mobile.startswith("0") and clean_mobile.isdigit():
                    clean_mobile = "0" + clean_mobile

                if not clean_mobile:
                    skipped += 1
                    continue

                name_str = str(name).strip()
                class_str = str(assigned_class).strip() if assigned_class else ""

                user = User.objects.filter(username=clean_mobile).first()
                if not user:
                    user = User.objects.create_user(
                        username=clean_mobile,
                        password="12345",
                    )

                obj, was_created = Teacher.objects.update_or_create(
                    name=name_str,
                    mobile=clean_mobile,
                    defaults={
                        "assigned_classes": class_str,
                        "user": user,
                        "employment_type": employment_type,
                    }
                )

                if obj.user_id != user.id:
                    obj.user = user
                    obj.save(update_fields=["user"])

                if was_created:
                    created += 1
                else:
                    updated += 1

            file_results.append({
                "filename": excel_file.name,
                "created": created,
                "updated": updated,
                "skipped": skipped,
            })

    return render(request, 'attendance/teacher_upload.html', {
        'file_results': file_results,
        'employment_type': employment_type,
        'type_label': type_label,
    })


@login_required
@user_passes_test(is_admin)
def teacher_delete(request, teacher_id):
    teacher = get_object_or_404(Teacher, id=teacher_id)
    if request.method == 'POST':
        linked_user = teacher.user
        teacher.delete()
        if linked_user:
            linked_user.delete()
        return redirect('teacher_list')
    return render(request, 'attendance/teacher_confirm_delete.html', {'teacher': teacher})


@login_required
@user_passes_test(is_admin)
def mark_teacher_attendance(request):
    today = timezone.now().date()
    date_str = today.strftime("%d-%b-%y")
    saved = False
    sms_sent_count = 0
    sms_warning = None

    already_marked = TeacherAttendance.objects.filter(date=today).exists()

    if request.method == 'POST' and not already_marked:
        teachers = Teacher.objects.all().order_by('id')
        sms_failed = []
        for teacher in teachers:
            status = request.POST.get(f'tatt_{teacher.id}', 'present')
            is_present = status == 'present'

            TeacherAttendance.objects.update_or_create(
                teacher=teacher,
                date=today,
                defaults={'is_present': is_present}
            )

            if not is_present:
                if teacher.mobile:
                    message = build_teacher_absent_message(teacher.name, date_str)
                    success, resp = send_sms(teacher.mobile, message)
                    if success:
                        sms_sent_count += 1
                    else:
                        sms_failed.append(teacher.name)
                else:
                    sms_failed.append(f"{teacher.name} (no phone)")

        saved = True
        already_marked = True
        if sms_failed:
            sms_warning = f"SMS could not be sent to: {', '.join(sms_failed)}"

    teachers = Teacher.objects.all().order_by('id')
    attendance_map = {
        a.teacher_id: a.is_present
        for a in TeacherAttendance.objects.filter(date=today)
    }
    teacher_rows = [
        {
            'teacher': t,
            'is_present': attendance_map.get(t.id, True),
        }
        for t in teachers
    ]

    full_time_rows = [r for r in teacher_rows if r['teacher'].employment_type == Teacher.EMPLOYMENT_FULL]
    part_time_rows = [r for r in teacher_rows if r['teacher'].employment_type == Teacher.EMPLOYMENT_PART]

    if already_marked:
        present_count = TeacherAttendance.objects.filter(date=today, is_present=True).count()
        absent_count = TeacherAttendance.objects.filter(date=today, is_present=False).count()
    else:
        present_count = 0
        absent_count = 0

    return render(request, 'attendance/mark_teacher_attendance.html', {
        'teachers': teachers,
        'teacher_rows': teacher_rows,
        'full_time_rows': full_time_rows,
        'part_time_rows': part_time_rows,
        'today': today,
        'saved': saved,
        'sms_sent_count': sms_sent_count,
        'sms_warning': sms_warning,
        'already_marked': already_marked,
        'present_count': present_count,
        'absent_count': absent_count,
        'total_teachers': teachers.count(),
    })

@login_required
@user_passes_test(is_admin)
def teacher_attendance_history(request):
    month_str = request.GET.get('month', '') or timezone.now().strftime('%Y-%m')
    try:
        year, month = map(int, month_str.split('-'))
    except ValueError:
        now = timezone.now()
        year, month = now.year, now.month
        month_str = f"{year:04d}-{month:02d}"

    teachers = Teacher.objects.all().order_by('id')
    records = []
    for teacher in teachers:
        qs = TeacherAttendance.objects.filter(teacher=teacher, date__year=year, date__month=month)
        present = qs.filter(is_present=True).count()
        absent = qs.filter(is_present=False).count()
        records.append({'teacher': teacher, 'present': present, 'absent': absent})

    return render(request, 'attendance/teacher_attendance_history.html', {
        'records': records,
        'month_str': month_str,
    })


@login_required
@user_passes_test(is_admin)
def export_teacher_attendance(request):
    month_str = request.GET.get('month', '') or timezone.now().strftime('%Y-%m')
    try:
        year, month = map(int, month_str.split('-'))
    except ValueError:
        now = timezone.now()
        year, month = now.year, now.month
        month_str = f"{year:04d}-{month:02d}"

    teachers = Teacher.objects.all().order_by('id')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Teacher Attendance {month_str}"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3D32", end_color="1E3D32", fill_type="solid")
    thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    headers = ["ID", "Name", "Mobile", "Present", "Absent"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin

    for row_num, teacher in enumerate(teachers, 2):
        qs = TeacherAttendance.objects.filter(teacher=teacher, date__year=year, date__month=month)
        present = qs.filter(is_present=True).count()
        absent = qs.filter(is_present=False).count()

        values = [
            teacher.id,
            teacher.name,
            teacher.mobile or "",
            present,
            absent,
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.border = thin
            cell.alignment = Alignment(horizontal="center" if col != 2 else "left")

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12

    filename = f"Teacher_Attendance_{month_str}.xlsx"
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
@user_passes_test(is_admin)
def attendance_history(request):
    class_names = Student.objects.values_list('class_name', flat=True).distinct().order_by('class_name')
    class_filter = request.GET.get('class', '').strip()
    date_filter = request.GET.get('date', '') or timezone.now().date().isoformat()

    all_classes = [{'name': c, 'is_selected': (str(c) == class_filter)} for c in class_names]

    records = []
    if class_filter:
        students = Student.objects.filter(class_name=class_filter).order_by('roll_no')
        attendance_map = {
            a.student_id: a.is_present
            for a in Attendance.objects.filter(student__class_name=class_filter, date=date_filter)
        }
        for student in students:
            status = attendance_map.get(student.id)
            records.append({
                'student': student,
                'status': 'present' if status is True else ('absent' if status is False else 'not_marked'),
            })

    present_count = sum(1 for r in records if r['status'] == 'present')
    absent_count = sum(1 for r in records if r['status'] == 'absent')

    return render(request, 'attendance/attendance_history.html', {
        'all_classes': all_classes,
        'class_filter': class_filter,
        'date_filter': date_filter,
        'records': records,
        'present_count': present_count,
        'absent_count': absent_count,
    })


@login_required
@user_passes_test(is_admin)
def export_attendance(request):
    class_filter = request.GET.get('class', '').strip()
    date_filter = request.GET.get('date', '') or timezone.now().date().isoformat()

    if not class_filter:
        return redirect('attendance_history')

    students = Student.objects.filter(class_name=class_filter).order_by('roll_no')
    attendance_map = {
        a.student_id: a.is_present
        for a in Attendance.objects.filter(student__class_name=class_filter, date=date_filter)
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Class {class_filter}"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3D32", end_color="1E3D32", fill_type="solid")
    present_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    absent_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    headers = ["Roll", "Name", "Class", "Section", "Mobile", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin

    for row_num, student in enumerate(students, 2):
        status_bool = attendance_map.get(student.id)
        if status_bool is True:
            status = "Present"
            fill = present_fill
        elif status_bool is False:
            status = "Absent"
            fill = absent_fill
        else:
            status = "Not Marked"
            fill = None

        values = [
            student.roll_no,
            student.name,
            student.class_name,
            student.section,
            student.parent_mobile or "",
            status,
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.border = thin
            cell.alignment = Alignment(horizontal="center" if col != 2 else "left")
            if fill and col == 6:
                cell.fill = fill

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12

    filename = f"Attendance_Class_{class_filter}_{date_filter}.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
